from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import os

def match_photos_to_system_ids(excel_file, photo_folder, output_file):
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    new_wb = Workbook()
    new_ws = new_wb.active
    
    image_files = os.listdir(photo_folder)
    
    header_row = ws[1]
    for cell in header_row:
        new_ws[cell.coordinate].value = cell.value
    
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True): 
        system_id = row[1]  
        image_filename = f"{system_id}.jpg" 
        if image_filename in image_files:
            photo_path = os.path.join(photo_folder, image_filename)
            photo_column_index = 11  
            photo_column = new_ws.cell(row=row[0], column=photo_column_index) 
            
            img = Image(photo_path)
            
            img.width = 500
            img.height = 500
            
            new_ws.add_image(img, photo_column.coordinate)
    
    new_wb.save(output_file)

# path to excel file, photos folder and output file
excel_file_path = "employees.xlsx"
photo_folder_path = "/Photos"
output_file_path = "employees_with_photos.xlsx"
#calling the function to host file, host folder, output file
match_photos_to_system_ids(excel_file_path, photo_folder_path, output_file_path)
